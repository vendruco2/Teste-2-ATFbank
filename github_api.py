import requests
import openpyxl

response = requests.get('https://api.github.com/repositories')

if response.status_code == 200:
    data = response.json()
    total_repos = len(data)
    without_json = 0
    for repo in data:
        description = repo["description"]
        if description is not None and "json" not in description.lower():
            without_json += 1

    # Cria um novo arquivo Excel e adiciona os dados
    wb = openpyxl.Workbook()
    ws = wb.active
    ws['A1'] = 'Total de repositórios'
    ws['B1'] = 'Repositórios sem "JSON" na descrição'
    ws['C1'] = 'Repositórios com "JSON" na descrição'
    ws['A2'] = total_repos
    ws['B2'] = without_json
    ws['C2'] = total_repos - without_json

    # Salva o arquivo Excel
    wb.save('github_repos.xlsx')

    print('Arquivo Excel criado com sucesso!')
else:
    print('Erro ao fazer solicitação: ', response.status_code)
